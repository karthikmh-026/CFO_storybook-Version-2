import os
import openpyxl

target_str = "1965"
dir_path = r"c:\Users\ajala\Downloads\CFO PITTI"
for root, dirs, files in os.walk(dir_path):
    if any(p in root for p in ['venv', 'node_modules', '.git']):
        continue
    for f in files:
        fp = os.path.join(root, f)
        if f.endswith(('.py', '.txt', '.json', '.js', '.jsx', '.html', '.css', '.bat', '.ps1', '.docx', '.csv')):
            try:
                with open(fp, 'r', errors='ignore') as file:
                    for i, line in enumerate(file, 1):
                        if target_str in line:
                            print(f"Text match: {fp}:{i}: {line.strip()[:150]}")
            except Exception as e:
                pass
        elif f.endswith(('.xlsx', '.xlsm')):
            try:
                wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        for c_idx, val in enumerate(row, 1):
                            if val is not None and target_str in str(val):
                                print(f"Excel match: {fp} | Sheet: {sheet} | Cell ({r_idx},{c_idx}): {val}")
            except Exception as e:
                pass
print("Search complete.")
