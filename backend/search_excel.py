import openpyxl

files = [
    r"c:\Users\ajala\Downloads\CFO_KPI_Logic.xlsx",
    r"c:\Users\ajala\Downloads\CFO_Storybook_KPI_Dictionary.xlsx"
]

for fp in files:
    print(f"\n=== Searching in {fp} ===")
    try:
        wb = openpyxl.load_workbook(fp, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                for c_idx, val in enumerate(row, 1):
                    if val is not None and ('1965' in str(val) or '1856' in str(val)):
                        print(f"Match: Sheet={sheet} | Cell ({r_idx},{c_idx}) = {val}")
    except Exception as e:
        print("Error reading:", e)
