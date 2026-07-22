import openpyxl
import os

files = [
    r"C:\Users\ajala\Downloads\SAP LOGIC.xlsx",
    r"C:\Users\ajala\Downloads\SAP_ALL_16_Dashboards_KPI_Logic.xlsx",
    r"C:\Users\ajala\Downloads\1000TB.xlsx",
    r"C:\Users\ajala\Downloads\2000TB.XLSX",
    r"C:\Users\ajala\Downloads\4000tb.xlsx"
]

targets = [1965.44, 1856.82, 320.19, 319.91]

for filepath in files:
    if not os.path.exists(filepath):
        print(f"Skipping {filepath} - does not exist")
        continue
    print(f"\nSearching in {filepath}...")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                for c_idx, val in enumerate(row, 1):
                    if val is not None:
                        # try converting to float
                        try:
                            num = float(val)
                            # check if close to any target (in Crore)
                            for t in targets:
                                if abs(num - t) < 0.5:
                                    print(f"  [CR] Sheet: {sheet} | Cell ({r_idx},{c_idx}) = {num} (target {t})")
                                # also check raw value if target is in rupees (t * 10^7)
                                if abs(num - t * 10_000_000.0) < 5_000_000.0:
                                    print(f"  [RAW] Sheet: {sheet} | Cell ({r_idx},{c_idx}) = {num} (target {t} Cr)")
                        except ValueError:
                            # check as string
                            val_str = str(val)
                            for t in targets:
                                if str(t) in val_str or str(int(t)) in val_str:
                                    print(f"  [STR] Sheet: {sheet} | Cell ({r_idx},{c_idx}) = {val} (target {t})")
    except Exception as e:
        print(f"  Error reading {filepath}: {e}")
