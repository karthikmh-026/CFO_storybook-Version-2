import os
import openpyxl

search_dir = r"C:\Users\ajala\Downloads"
targets = ["1965", "1856", "320.19", "319.91", "108.62"]

print("Starting verbose search...")
for root, dirs, files in os.walk(search_dir):
    # skip system/venv folders
    if any(p in root for p in ['venv', 'node_modules', '.git', 'Anaconda3', 'SAPGUI750']):
        continue
    for f in files:
        fp = os.path.join(root, f)
        ext = f.split('.')[-1].lower()
        if ext in ['txt', 'py', 'json', 'csv', 'docx']:
            try:
                with open(fp, 'r', encoding='utf-8', errors='ignore') as file:
                    content = file.read()
                    for t in targets:
                        if t in content:
                            print(f"Text Match in {fp}: found '{t}'")
            except Exception as e:
                pass
        elif ext in ['xlsx', 'xlsm']:
            try:
                # read excel sheets
                wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
                for s in wb.sheetnames:
                    ws = wb[s]
                    for r_idx, r in enumerate(ws.iter_rows(values_only=True), 1):
                        for c_idx, val in enumerate(r, 1):
                            if val is not None:
                                val_str = str(val)
                                for t in targets:
                                    if t in val_str:
                                        print(f"Excel Match in {fp} | Sheet: {s} | Cell ({r_idx},{c_idx}) = {val}")
            except Exception as e:
                pass
print("Verbose search complete.")
