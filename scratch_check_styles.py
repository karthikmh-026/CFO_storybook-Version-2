import openpyxl
from openpyxl.styles import PatternFill, Font

def main():
    wb = openpyxl.load_workbook("CFO_Storybook_KPI_Dictionary.xlsx")
    ws = wb["KPI Dictionary"]
    
    # We will inspect a few rows with different statuses
    # Row 2 is "Yes", Row 8 is "Partial", Row 11 is "No"
    for r in [2, 8, 11]:
        cell = ws.cell(row=r, column=4) # Column D (Status)
        val = cell.value
        fill = cell.fill
        font = cell.font
        
        print(f"Row {r}: Status={val}")
        if isinstance(fill, PatternFill):
            print(f"  Fill: type={fill.fill_type}, fgColor={fill.fgColor.rgb if fill.fgColor else 'None'}")
        if isinstance(font, Font):
            print(f"  Font: name={font.name}, size={font.size}, bold={font.bold}, color={font.color.rgb if font.color else 'None'}")

if __name__ == '__main__':
    main()
